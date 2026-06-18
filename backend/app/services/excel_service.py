from __future__ import annotations

import io
import json
from datetime import datetime
from typing import Any

from openpyxl import Workbook, load_workbook

from app.schemas.record import BatchExcelImportResponse, BatchExcelRow


class ExcelService:
    def parse_batch_import(self, content: bytes) -> BatchExcelImportResponse:
        workbook = load_workbook(filename=io.BytesIO(content), data_only=True)
        sheet = workbook.active

        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header_row:
            raise ValueError("Excel 文件缺少表头")

        headers = [str(cell).strip() if cell is not None else "" for cell in header_row]
        if "user_input" not in headers:
            raise ValueError("Excel 文件必须包含 user_input 列")

        user_input_index = headers.index("user_input")

        items: list[BatchExcelRow] = []
        for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            user_input = str(row[user_input_index]).strip() if row[user_input_index] is not None else ""
            if not user_input:
                continue

            items.append(
                BatchExcelRow(
                    row_number=row_number,
                    user_input=user_input,
                )
            )

        return BatchExcelImportResponse(items=items, total=len(items))

    def export_records_excel(self, records: list[dict[str, Any]]) -> bytes:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "consultation_records"

        headers = [
            "id",
            "created_at",
            "selected_persona_name",
            "rag_ready",
            "sample_reason",
            "sample_tags_json",
            "planner_labels_json",
            "risk_assessment_json",
            "evaluation_total",
            "evaluation_problem_risk_recognition",
            "evaluation_emotional_response_moderation",
            "evaluation_cognitive_reframing",
            "evaluation_advice_effectiveness",
            "evaluation_value_guidance_safety",
            "evaluation_has_safety_issue",
            "evaluation_has_safety_advice",
            "evaluation_safety_advice_effective",
            "evaluation_has_safety_issue_note",
            "evaluation_has_safety_advice_note",
            "evaluation_safety_advice_effective_note",
            "evaluation_json",
            "user_input",
            "ai_selected_raw_response",
            "expert_polished_response",
            "expert_annotation",
            "sample_snapshot_json",
            "selected_style_config_json",
            "planner_output_json",
            "draft_candidates_json",
        ]
        sheet.append(headers)

        for record in records:
            created_at = record.get("created_at")
            if isinstance(created_at, datetime):
                created_at_value = created_at.isoformat()
            else:
                created_at_value = str(created_at or "")
            evaluation = record.get("evaluation_json", {}) or {}
            scores = evaluation.get("scores", {}) or {}
            safety_checks = evaluation.get("safety_checks", {}) or {}
            safety_notes = evaluation.get("safety_notes", {}) or {}

            sheet.append(
                [
                    record.get("id", ""),
                    created_at_value,
                    record.get("selected_persona_name", ""),
                    record.get("rag_ready", ""),
                    record.get("sample_reason", ""),
                    json.dumps(record.get("sample_tags_json", {}), ensure_ascii=False),
                    json.dumps(record.get("planner_labels_json", {}), ensure_ascii=False),
                    json.dumps(record.get("risk_assessment_json", {}), ensure_ascii=False),
                    evaluation.get("total_score", ""),
                    self._score_with_legacy(scores, "problem_risk_recognition"),
                    self._score_with_legacy(scores, "emotional_response_moderation"),
                    self._score_with_legacy(scores, "cognitive_reframing"),
                    self._score_with_legacy(scores, "advice_effectiveness"),
                    self._score_with_legacy(scores, "value_guidance_safety"),
                    self._format_yes_no(safety_checks.get("has_safety_issue")),
                    self._format_yes_no(safety_checks.get("has_safety_advice")),
                    self._format_yes_no(safety_checks.get("safety_advice_effective")),
                    safety_notes.get("has_safety_issue", ""),
                    safety_notes.get("has_safety_advice", ""),
                    safety_notes.get("safety_advice_effective", ""),
                    json.dumps(evaluation, ensure_ascii=False),
                    record.get("user_input", ""),
                    record.get("ai_selected_raw_response", ""),
                    record.get("expert_polished_response", ""),
                    record.get("expert_annotation", ""),
                    json.dumps(record.get("sample_snapshot_json", {}), ensure_ascii=False),
                    json.dumps(record.get("selected_style_config_json", {}), ensure_ascii=False),
                    json.dumps(record.get("planner_output_json", {}), ensure_ascii=False),
                    json.dumps(record.get("draft_candidates_json", []), ensure_ascii=False),
                ]
            )

        output = io.BytesIO()
        workbook.save(output)
        return output.getvalue()

    def export_batch_generation_excel(self, results: list[dict[str, Any]]) -> bytes:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "batch_generation_results"

        headers = [
            "row_number",
            "user_input",
            "selected_persona_names",
            "draft_count",
            "persona_name",
            "response",
            "planner_output_json",
        ]
        sheet.append(headers)

        for result in results:
            drafts = result.get("drafts", [])
            if not drafts:
                sheet.append(
                    [
                        result.get("row_number", ""),
                        result.get("user_input", ""),
                        ",".join(result.get("selected_persona_names", [])),
                        0,
                        "",
                        "",
                        "",
                    ]
                )
                continue

            for draft in drafts:
                sheet.append(
                    [
                        result.get("row_number", ""),
                        result.get("user_input", ""),
                        ",".join(result.get("selected_persona_names", [])),
                        result.get("draft_count", 0),
                        draft.get("persona_name", ""),
                        draft.get("response", ""),
                        json.dumps(draft.get("planner_output", {}), ensure_ascii=False),
                    ]
                )

        output = io.BytesIO()
        workbook.save(output)
        return output.getvalue()

    def export_reviewed_batch_excel(self, items: list[dict[str, Any]]) -> bytes:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "reviewed_batch_results"

        headers = [
            "row_number",
            "user_input",
            "selected_persona_name",
            "rag_ready",
            "sample_reason",
            "risk_assessment_json",
            "evaluation_total",
            "evaluation_problem_risk_recognition",
            "evaluation_emotional_response_moderation",
            "evaluation_cognitive_reframing",
            "evaluation_advice_effectiveness",
            "evaluation_value_guidance_safety",
            "evaluation_has_safety_issue",
            "evaluation_has_safety_advice",
            "evaluation_safety_advice_effective",
            "evaluation_has_safety_issue_note",
            "evaluation_has_safety_advice_note",
            "evaluation_safety_advice_effective_note",
            "evaluation_json",
            "final_response",
            "expert_annotation",
        ]
        sheet.append(headers)

        for item in items:
            evaluation = item.get("evaluation", {}) or {}
            scores = evaluation.get("scores", {}) or {}
            safety_checks = evaluation.get("safety_checks", {}) or {}
            safety_notes = evaluation.get("safety_notes", {}) or {}
            sheet.append(
                [
                    item.get("row_number", ""),
                    item.get("user_input", ""),
                    item.get("selected_persona_name", ""),
                    item.get("rag_ready", ""),
                    item.get("sample_reason", ""),
                    json.dumps(item.get("risk_assessment", {}), ensure_ascii=False),
                    evaluation.get("total_score", ""),
                    self._score_with_legacy(scores, "problem_risk_recognition"),
                    self._score_with_legacy(scores, "emotional_response_moderation"),
                    self._score_with_legacy(scores, "cognitive_reframing"),
                    self._score_with_legacy(scores, "advice_effectiveness"),
                    self._score_with_legacy(scores, "value_guidance_safety"),
                    self._format_yes_no(safety_checks.get("has_safety_issue")),
                    self._format_yes_no(safety_checks.get("has_safety_advice")),
                    self._format_yes_no(safety_checks.get("safety_advice_effective")),
                    safety_notes.get("has_safety_issue", ""),
                    safety_notes.get("has_safety_advice", ""),
                    safety_notes.get("safety_advice_effective", ""),
                    json.dumps(evaluation, ensure_ascii=False),
                    item.get("final_response", ""),
                    item.get("expert_annotation", ""),
                ]
            )

        output = io.BytesIO()
        workbook.save(output)
        return output.getvalue()

    @staticmethod
    def _score_with_legacy(scores: dict[str, Any], key: str) -> Any:
        legacy_keys = {
            "problem_risk_recognition": "intent_safety",
            "emotional_response_moderation": "authentic_empathy",
            "cognitive_reframing": "grounded_guidance",
        }
        return scores.get(key, scores.get(legacy_keys.get(key, ""), ""))

    @staticmethod
    def _format_yes_no(value: Any) -> str:
        if value is True:
            return "是"
        if value is False:
            return "否"
        return ""
