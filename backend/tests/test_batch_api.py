from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

from app.core.config import Settings
from app.main import create_app


def build_test_client(tmp_path):
    settings = Settings(
        database_url=f"sqlite:///{tmp_path / 'batch.db'}",
        mock_llm=True,
    )
    app = create_app(settings)
    return TestClient(app)


def make_test_excel() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["user_input", "selected_persona_names"])
    sheet.append(["我最近压力很大，想找人说说。", "温暖倾听者,理性破局教练"])
    sheet.append(["我总觉得自己快扛不住了。", "启发故事导师"])

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def test_batch_import_excel(tmp_path):
    client = build_test_client(tmp_path)
    excel_bytes = make_test_excel()

    response = client.post(
        "/api/v1/batch/import",
        files={"file": ("batch.xlsx", excel_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["total"] == 2
    assert payload["items"][0]["selected_persona_names"] == []


def test_export_records_excel(tmp_path):
    client = build_test_client(tmp_path)
    payload = {
        "user_input": "我最近特别累，不知道该怎么办。",
        "selected_persona_name": "温暖倾听者",
        "selected_style_config": {"persona_name": "温暖倾听者"},
        "planner_output": {"generation_plan": "先接住情绪，再给小动作"},
        "draft_candidates": [
            {
                "persona_name": "温暖倾听者",
                "style_config": {"persona_name": "温暖倾听者"},
                "planner_output": {"generation_plan": "test"},
                "response": "你好，我在这里。",
                "raw_response": "",
            }
        ],
        "ai_selected_raw_response": "你好，我在这里。",
        "expert_polished_response": "你好，我认真看见了你的疲惫。",
        "expert_annotation": "补充了具体求助对象。",
        "rag_ready": "approved",
        "sample_reason": "这条适合作为 RAG 样本。",
        "sample_snapshot": {"final_response": "你好，我认真看见了你的疲惫。"},
    }
    client.post("/api/v1/records", json=payload)

    response = client.get("/api/v1/batch/records/export")
    assert response.status_code == 200

    workbook = load_workbook(BytesIO(response.content))
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    assert rows[0][3] == "rag_ready"
    assert rows[1][3] == payload["rag_ready"]
    assert rows[0][4] == "sample_reason"
    assert rows[1][4] == payload["sample_reason"]
    assert rows[0][5] == "sample_tags_json"
    assert rows[0][6] == "planner_labels_json"
    assert rows[0][10] == "expert_annotation"
    assert rows[1][10] == payload["expert_annotation"]
