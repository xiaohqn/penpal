"""
输入：
- 指向临时 SQLite 文件的测试配置。
- 针对安全回复记录接口构造的创建、列表、详情、删除和导出请求。
输出：
- 验证安全回复记录接口的保存、分页、详情读取、删除和导出行为是否符合预期。
作用：
- 这个测试文件覆盖安全回复样本库的核心后端链路，避免后续改动破坏落库、查询与导出能力。
"""
from fastapi.testclient import TestClient
from io import BytesIO

from openpyxl import load_workbook

from app.core.config import Settings
from app.main import create_app


def build_test_client(tmp_path):
    """
    输入：
    - tmp_path：pytest 提供的临时目录。
    输出：
    - 返回一个使用临时 SQLite 数据库的 FastAPI 测试客户端。
    作用：
    - 为安全回复记录接口测试提供彼此隔离的运行环境。
    """

    settings = Settings(
        database_url=f"sqlite:///{tmp_path / 'safety_records.db'}",
        mock_llm=True,
    )
    app = create_app(settings)
    return TestClient(app)


def test_safety_record_create_and_fetch(tmp_path):
    """
    输入：
    - 临时测试数据库目录。
    输出：
    - 断言创建、列表和详情接口都能正确读写安全回复记录。
    作用：
    - 验证安全回复记录样本库的基本增查链路可用。
    """

    client = build_test_client(tmp_path)

    payload = {
        "user_input": "我最近总觉得自己快坚持不下去了。",
        "risk_labels": ["自杀倾向"],
        "corrected_risk_labels": ["自杀倾向", "其它风险"],
        "risk_reason": "来信中出现了明确的轻生表达。",
        "ai_safe_response": "谢谢你愿意把这些说出来，你现在的安全最重要。",
        "expert_polished_response": "谢谢你把这些告诉我。你现在的安全最重要，我们先一起确保你身边有人支持。",
        "selected_response_source": "api",
        "selected_response_source_label": "API 安全回复",
        "safe_response_candidates": [
            {
                "source": "api",
                "source_label": "API 安全回复",
                "intent": "先稳住安全，再推动现实求助。",
                "safe_response": "谢谢你愿意把这些说出来，你现在的安全最重要。",
                "safe_highlight_segments": ["你现在的安全最重要"],
                "safe_highlight_source": "llm",
            }
        ],
        "expert_annotation": "我补强了现实求助对象和一句更能落地的建议。",
        "sample_snapshot": {
            "selected_response_source": "api",
            "response_versions": [
                {
                    "version_index": 0,
                    "label": "专家当前版本",
                    "response": "谢谢你把这些告诉我。你现在的安全最重要，我们先一起确保你身边有人支持。",
                }
            ],
        },
        "source_annotations": [
            {
                "id": "annotation-1",
                "start": 0,
                "end": 7,
                "quote": "你现在的安全最重要",
                "note": "这里需要保留，并补一句现实支持对象。",
                "color": "amber",
            }
        ],
        "response_versions": [
            {
                "version_index": 0,
                "label": "专家当前版本",
                "response": "谢谢你把这些告诉我。你现在的安全最重要，我们先一起确保你身边有人支持。",
                "selected_response_source": "api",
                "selected_response_source_label": "API 安全回复",
                "created_at": "2026-05-28T10:00:00+00:00",
                "source": "manual",
                "expert_annotation": "我补强了现实求助对象和一句更能落地的建议。",
                "source_annotations": [],
            }
        ],
    }

    create_response = client.post("/api/v1/safety-records", json=payload)
    assert create_response.status_code == 201
    record_id = create_response.json()["id"]

    list_response = client.get("/api/v1/safety-records")
    assert list_response.status_code == 200
    list_payload = list_response.json()
    assert list_payload["total"] == 1
    assert list_payload["items"][0]["style_name"] == "安全"

    detail_response = client.get(f"/api/v1/safety-records/{record_id}")
    assert detail_response.status_code == 200
    detail_payload = detail_response.json()
    assert detail_payload["risk_labels_json"] == payload["risk_labels"]
    assert detail_payload["corrected_risk_labels_json"] == payload["corrected_risk_labels"]
    assert detail_payload["expert_polished_response"] == payload["expert_polished_response"]
    assert detail_payload["selected_response_source"] == payload["selected_response_source"]
    assert detail_payload["selected_response_source_label"] == payload["selected_response_source_label"]
    assert detail_payload["expert_annotation"] == payload["expert_annotation"]
    assert detail_payload["source_annotations_json"][0]["note"] == payload["source_annotations"][0]["note"]
    assert detail_payload["response_versions_json"][0]["label"] == payload["response_versions"][0]["label"]
    assert detail_payload["safe_response_candidates_json"][0]["source"] == payload["safe_response_candidates"][0]["source"]


def test_safety_record_delete(tmp_path):
    """
    输入：
    - 临时测试数据库目录。
    输出：
    - 断言删除接口会真正移除记录，并在后续查询时返回 404。
    作用：
    - 验证安全回复记录历史页的删除按钮背后具备可靠的后端删除能力。
    """

    client = build_test_client(tmp_path)

    payload = {
        "user_input": "我现在真的不知道还能不能撑下去。",
        "risk_labels": ["自杀倾向"],
        "corrected_risk_labels": ["自杀倾向"],
        "risk_reason": "来信包含明显的绝望和轻生倾向表达。",
        "ai_safe_response": "谢谢你告诉我这些，我们先把你的安全放在第一位。",
        "expert_polished_response": "谢谢你把这些说出来。你现在的安全最重要，我们先一起找到能立刻支持你的人。",
    }

    create_response = client.post("/api/v1/safety-records", json=payload)
    assert create_response.status_code == 201
    record_id = create_response.json()["id"]

    delete_response = client.delete(f"/api/v1/safety-records/{record_id}")
    assert delete_response.status_code == 204
    assert delete_response.text == ""

    detail_response = client.get(f"/api/v1/safety-records/{record_id}")
    assert detail_response.status_code == 404

    list_response = client.get("/api/v1/safety-records")
    assert list_response.status_code == 200
    assert list_response.json()["total"] == 0


def test_safety_record_export_excel(tmp_path):
    """
    输入：
    - 临时测试数据库目录。
    输出：
    - 断言安全回复记录导出接口会返回包含关键字段的 Excel 文件。
    作用：
    - 验证历史页新增的“导出安全回复记录”按钮背后具备稳定的后端文件导出能力。
    """

    client = build_test_client(tmp_path)

    payload = {
        "user_input": "我最近总觉得自己快坚持不下去了。",
        "risk_labels": ["自杀倾向"],
        "corrected_risk_labels": ["自杀倾向", "其它风险"],
        "risk_reason": "来信中出现了明确的轻生表达。",
        "ai_safe_response": "谢谢你愿意把这些说出来，你现在的安全最重要。",
        "expert_polished_response": "谢谢你把这些告诉我。你现在的安全最重要，我们先一起确保你身边有人支持。",
    }

    create_response = client.post("/api/v1/safety-records", json=payload)
    assert create_response.status_code == 201

    export_response = client.get("/api/v1/safety-records/export")
    assert export_response.status_code == 200

    workbook = load_workbook(BytesIO(export_response.content))
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))

    assert rows[0][2] == "style_name"
    assert rows[0][3] == "selected_response_source_label"
    assert rows[0][4] == "risk_labels_json"
    assert rows[0][5] == "corrected_risk_labels_json"
    assert rows[0][8] == "ai_safe_response"
    assert rows[0][10] == "expert_annotation"
    assert rows[0][11] == "source_annotations_json"
    assert rows[0][12] == "response_versions_json"
    assert rows[1][2] == "安全"
    assert "自杀倾向" in str(rows[1][4])
    assert rows[1][8] == payload["ai_safe_response"]
    assert rows[1][9] == payload["expert_polished_response"]
